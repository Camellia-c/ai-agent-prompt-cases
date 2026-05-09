"""
Calculate student renewal risk from an Excel file.

Usage:
python calculate_renewal_risk.py input.xlsx output.xlsx report.md
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook, Workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl"
    ) from exc


REQUIRED_COLUMNS = [
    "student_name",
    "attendance_rate",
    "homework_rate",
    "interaction_score",
    "parent_reply_rate",
    "renewal_intent",
    "complaint_count",
    "last_contact_days",
]


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_rate(value: Any) -> float:
    number = to_float(value)
    if number < 0:
        return 0.0
    if number > 100:
        return 100.0
    return number


def renewal_intent_risk(value: Any) -> float:
    text = str(value or "unknown").strip().lower()
    mapping = {
        "high": 10,
        "medium": 40,
        "low": 80,
        "unknown": 60,
        "高": 10,
        "中": 40,
        "低": 80,
        "未知": 60,
    }
    return float(mapping.get(text, 60))


def complaint_risk(value: Any) -> float:
    count = int(to_float(value))
    if count >= 3:
        return 100.0
    if count == 2:
        return 70.0
    if count == 1:
        return 40.0
    return 0.0


def contact_risk(value: Any) -> float:
    days = to_float(value)
    if days >= 14:
        return 100.0
    if days >= 7:
        return 60.0
    return 20.0


def calculate_risk_score(row: dict[str, Any]) -> float:
    attendance = normalize_rate(row["attendance_rate"])
    homework = normalize_rate(row["homework_rate"])
    interaction = normalize_rate(row["interaction_score"])
    parent_reply = normalize_rate(row["parent_reply_rate"])

    score = (
        (100 - attendance) * 0.20
        + (100 - homework) * 0.20
        + (100 - interaction) * 0.15
        + (100 - parent_reply) * 0.15
        + renewal_intent_risk(row["renewal_intent"]) * 0.15
        + complaint_risk(row["complaint_count"]) * 0.10
        + contact_risk(row["last_contact_days"]) * 0.05
    )
    return round(score, 2)


def classify_risk(score: float) -> str:
    if score >= 70:
        return "High"
    if score >= 40:
        return "Medium"
    return "Low"


def follow_up_action(level: str) -> str:
    if level == "High":
        return "Call parent within 24 hours and provide a personalized learning review."
    if level == "Medium":
        return "Send private message within 48 hours and reinforce course value."
    return "Maintain regular follow-up and encourage continued learning."


def read_excel(input_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_to_index = {
        str(header).strip(): idx
        for idx, header in enumerate(headers)
        if header
    }

    missing = [col for col in REQUIRED_COLUMNS if col not in header_to_index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        record = {col: row[header_to_index[col]] for col in REQUIRED_COLUMNS}
        records.append(record)

    return records


def write_output(records: list[dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Renewal Risk"

    headers = REQUIRED_COLUMNS + ["risk_score", "risk_level", "follow_up_action"]
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(output_path)


def write_report(records: list[dict[str, Any]], report_path: Path) -> None:
    total = len(records)
    high = sum(1 for r in records if r["risk_level"] == "High")
    medium = sum(1 for r in records if r["risk_level"] == "Medium")
    low = sum(1 for r in records if r["risk_level"] == "Low")

    top_high_risk = sorted(records, key=lambda r: r["risk_score"], reverse=True)[:5]

    lines = [
        "# Renewal Risk Analysis Report",
        "",
        "## Summary",
        "",
        f"- Total students: {total}",
        f"- High risk: {high}",
        f"- Medium risk: {medium}",
        f"- Low risk: {low}",
        "",
        "## Top Risk Students",
        "",
        "| Student | Risk Score | Risk Level | Follow-up Action |",
        "|---|---:|---|---|",
    ]

    for record in top_high_risk:
        lines.append(
            f"| {record['student_name']} | {record['risk_score']} | "
            f"{record['risk_level']} | {record['follow_up_action']} |"
        )

    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- Risk score is calculated from attendance, homework completion, interaction, parent reply rate, renewal intent, complaints, and days since last contact.",
            "- Higher score means higher renewal risk.",
            "- This result is for operational prioritization and should be combined with real parent communication context.",
        ]
    )

    report_path.write_text("\n".join(lines), encoding="utf-8")


def analyze(input_path: Path, output_path: Path, report_path: Path) -> None:
    records = read_excel(input_path)

    for record in records:
        score = calculate_risk_score(record)
        level = classify_risk(score)
        record["risk_score"] = score
        record["risk_level"] = level
        record["follow_up_action"] = follow_up_action(level)

    write_output(records, output_path)
    write_report(records, report_path)


def main() -> None:
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: python calculate_renewal_risk.py input.xlsx output.xlsx report.md"
        )

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    report_path = Path(sys.argv[3])

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    analyze(input_path, output_path, report_path)

    print(f"Output Excel created: {output_path}")
    print(f"Markdown report created: {report_path}")


if __name__ == "__main__":
    main()
